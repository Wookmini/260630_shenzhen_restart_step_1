import os
import json
import openpyxl
import difflib

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "영수증 보관소", "정산내역_YYYY-MM(양식).xlsx")
MAPPING_JSON_PATH = os.path.join(BASE_DIR, "data", "account_mapping.json")

def extract_backdata_from_excel():
    """엑셀 양식의 숨김 시트들에서 과거 정산 내역(Backdata)을 추출하여 JSON으로 저장"""
    if not os.path.exists(TEMPLATE_PATH):
        return {}

    wb = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
    mapping_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state == 'hidden':
            if sheet_name == '세부내역':
                for row in ws.iter_rows(min_row=3, max_row=500, values_only=True):
                    desc = row[1]
                    major = row[3]
                    minor = row[5]
                    if desc and isinstance(desc, str) and major:
                        mapping_data.append({
                            "description": desc.strip(),
                            "major": major.strip() if isinstance(major, str) else major,
                            "minor": minor.strip() if isinstance(minor, str) else minor
                        })
            elif '업무추진비' in sheet_name:
                for row in ws.iter_rows(min_row=3, max_row=500, values_only=True):
                    desc = row[3]  # 사용처
                    major = "업무추진비"
                    purpose = row[6] # 사용목적
                    if desc and isinstance(desc, str):
                        # 사용목적이 있으면 그걸 소계정이나 내역의 힌트로 쓸 수 있지만, 양식 상 소계정은 보통 없음
                        mapping_data.append({
                            "description": desc.strip(),
                            "major": major,
                            "minor": None
                        })

    # 중복 제거
    unique_mapping = []
    seen = set()
    for item in mapping_data:
        key = (item['description'], item['major'], item['minor'])
        if key not in seen:
            seen.add(key)
            unique_mapping.append(item)

    os.makedirs(os.path.dirname(MAPPING_JSON_PATH), exist_ok=True)
    with open(MAPPING_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(unique_mapping, f, indent=2, ensure_ascii=False)
    
    return unique_mapping

def load_mapping_data():
    if not os.path.exists(MAPPING_JSON_PATH):
        return extract_backdata_from_excel()
    with open(MAPPING_JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

# 기본 수동 키워드 사전 (Backdata에서 유추하기 어려운 것들 보완)
MANUAL_KEYWORDS = {
    "식대": {"major": "업무추진비", "minor": None, "desc": "식대"},
    "식당": {"major": "업무추진비", "minor": None, "desc": "심천 식당"},
    "coffee": {"major": "업무추진비", "minor": None, "desc": "Coffee Shop"},
    "커피": {"major": "업무추진비", "minor": None, "desc": "Coffee Shop"},
    "택시": {"major": "여비교통비", "minor": "교통비", "desc": "교통비"},
    "滴滴": {"major": "여비교통비", "minor": "교통비", "desc": "교통비"},
    "出租": {"major": "여비교통비", "minor": "교통비", "desc": "교통비"},
    "교통": {"major": "여비교통비", "minor": "교통비", "desc": "교통비"},
    "톨비": {"major": "여비교통비", "minor": "교통비", "desc": "톨비"},
    "高速": {"major": "여비교통비", "minor": "교통비", "desc": "톨비"},
    "公司": {"major": "여비교통비", "minor": "교통비", "desc": "교통비/톨비"},
    "限公司": {"major": "여비교통비", "minor": "교통비", "desc": "교통비/톨비"},
    "股份": {"major": "여비교통비", "minor": "교통비", "desc": "교통비/톨비"},
    "문구": {"major": "해외지사비", "minor": "기타집기", "desc": "사무용품"},
    "용지": {"major": "해외지사비", "minor": "사무실관리", "desc": "A4용지 및 생수 등"},
    "펜": {"major": "해외지사비", "minor": "기타집기", "desc": "사무용품"},
    "得力": {"major": "해외지사비", "minor": "기타집기", "desc": "사무용품"}, # 델리 (중국 문구브랜드)
    "화물운송료": {"major": "해외지사비", "minor": "운반비", "desc": "화물운송료"},
    "货拉拉": {"major": "해외지사비", "minor": "운반비", "desc": "화물운송료"},
}

LEARNING_MODEL_PATH = os.path.join(BASE_DIR, "data", "shenzhen_receipt_learning_model.json")
import re

def map_receipt_data(raw_description: str, seller: str = "", raw_text: str = "") -> dict:
    """
    OCR로 추출한 날것의 내역(raw_description)과 판매자(seller)를 바탕으로
    과거 엑셀 정산자료(Backdata) 및 키워드를 참조하여 대계정, 소계정, 표준 내역을 추론함.
    """
    if not raw_description:
        raw_description = ""
    if not seller:
        seller = ""
    if not raw_text:
        raw_text = ""
        
    combined_text = (raw_description + " " + seller).lower()
    
    # 0. 수동 키워드 룰셋 먼저 확인 (가장 높은 우선순위)
    for kw, mapping in MANUAL_KEYWORDS.items():
        if kw.lower() in combined_text:
            return {
                "is_mapped": True,
                "major": mapping["major"],
                "minor": mapping["minor"],
                "standard_desc": mapping["desc"]
            }

    # 1. 과거 5월 참고데이터 학습 모델(shenzhen_receipt_learning_model.json) 매칭 시도
    if raw_text and os.path.exists(LEARNING_MODEL_PATH):
        try:
            with open(LEARNING_MODEL_PATH, "r", encoding="utf-8") as f:
                learning_data = json.load(f)
            
            best_score = 0.0
            best_rule = None
            
            t1 = set(re.findall(r'[\u4e00-\u9fff\w]+', raw_text.lower()))
            if t1:
                for item in learning_data:
                    ocr_pages = item.get("ocr_pages", [])
                    rules = item.get("associated_rules", [])
                    
                    for p_idx, page_text in enumerate(ocr_pages):
                        t2 = set(re.findall(r'[\u4e00-\u9fff\w]+', page_text.lower()))
                        if not t2:
                            continue
                        intersection = t1.intersection(t2)
                        score = len(intersection) / min(len(t1), len(t2))
                        
                        if score > best_score:
                            best_score = score
                            # 1대1 매핑 또는 첫번째 룰 선택
                            if len(rules) == 1:
                                best_rule = rules[0]
                            elif p_idx < len(rules):
                                best_rule = rules[p_idx]
                            else:
                                best_rule = rules[0] if rules else None
            
            # 높은 유사도(60% 이상)일 때 매핑 신뢰
            if best_score > 0.6 and best_rule:
                major = best_rule.get("대계정")
                minor = best_rule.get("소계정")
                desc = best_rule.get("내역")
                
                # 대계정이 없더라도 학습 모델에 부합하는 정상이면 is_mapped=True 처리 (예: USD 환전)
                return {
                    "is_mapped": True,
                    "major": major,
                    "minor": minor,
                    "standard_desc": desc if desc else raw_description
                }
        except Exception as e:
            print(f"[Learning Model Match Error] {e}")

    # 2. Backdata(과거 내역) 텍스트 유사도 매칭 (difflib 활용)
    backdata = load_mapping_data()
    descriptions = [item['description'] for item in backdata]
    
    # 판매자나 내역과 과거 내역(description)이 비슷한지 확인
    matches = difflib.get_close_matches(raw_description, descriptions, n=1, cutoff=0.5)
    if not matches and seller:
        matches = difflib.get_close_matches(seller, descriptions, n=1, cutoff=0.5)
        
    if matches:
        best_match = matches[0]
        for item in backdata:
            if item['description'] == best_match:
                return {
                    "is_mapped": True,
                    "major": item['major'],
                    "minor": item['minor'],
                    "standard_desc": item['description']
                }

    # 3. 매핑 실패
    return {
        "is_mapped": False,
        "major": None,
        "minor": None,
        "standard_desc": raw_description if raw_description else seller
    }

if __name__ == "__main__":
    # Test extraction and mapping
    data = extract_backdata_from_excel()
    print(f"Extracted {len(data)} backdata rules.")
    
    # Test cases
    tests = ["得力彩色中性笔", "滴滴出行", "심천 한 식당", "알수없는 이상한 내역"]
    for t in tests:
        res = map_receipt_data(t)
        print(f"[{t}] -> {res}")

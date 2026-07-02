import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('영수증 보관소/2026-05/정산내역_2026-05_v3양식.xlsx', data_only=True)
sheet = wb.active

print("Rows in v3 Excel:")
# Headers are in some row, let's search for "Chen Feng Ju"
for row_idx in range(1, sheet.max_row + 1):
    val_e = sheet.cell(row=row_idx, column=5).value # Column E is assignee name? Or name column?
    # Let's print the entire row if it contains "Chen Feng Ju" or "Chen FengJu"
    row_vals = [sheet.cell(row=row_idx, column=col).value for col in range(1, 20)]
    if any(isinstance(v, str) and "Chen" in v for v in row_vals):
        print(f"Row {row_idx}: {row_vals}")

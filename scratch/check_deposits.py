import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('영수증 보관소/2026-05/정산내역_2026-05_v3양식.xlsx', data_only=True)
sheet = wb.active

print("Rows with non-empty deposits (Column J, K, L):")
for row_idx in range(1, sheet.max_row + 1):
    j_val = sheet.cell(row=row_idx, column=10).value
    k_val = sheet.cell(row=row_idx, column=11).value
    l_val = sheet.cell(row=row_idx, column=12).value
    
    if j_val is not None or k_val is not None or l_val is not None:
        row_vals = [sheet.cell(row=row_idx, column=col).value for col in range(1, 16)]
        print(f"Row {row_idx}: {row_vals}")

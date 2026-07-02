import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

template_path = r"C:\Users\20000177\Desktop\Wooktigravity\260630_shenzhen_restart_step_1\영수증 보관소\2026-05\정산내역_2026-05_v4.xlsx"
wb = openpyxl.load_workbook(template_path, data_only=True)
sheet = wb.active

print("Ev_No | Date (C) | WDate (B) | Person (F)")
for r_idx in range(21, 36):
    ev = sheet.cell(row=r_idx, column=7).value
    date = sheet.cell(row=r_idx, column=3).value
    wdate = sheet.cell(row=r_idx, column=2).value
    person = sheet.cell(row=r_idx, column=6).value
    print(f"{ev} | {date} | {wdate} | {person}")

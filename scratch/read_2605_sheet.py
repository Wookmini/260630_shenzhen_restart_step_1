import openpyxl

wb = openpyxl.load_workbook('과거 참고데이터/26년 5월/심천지사 전도금 정산 양식 변경 요청_260525_재정비.xlsx', data_only=True)
ws = wb['26.05']

print("Row 1-5:")
for r in range(1, 10):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
    print(f"Row {r}: {row_vals}")

print("\nRows with E-column values:")
for r in range(1, ws.max_row + 1):
    e_val = ws.cell(row=r, column=5).value  # Column E is 5
    if e_val is not None:
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 12)]
        print(f"Row {r}: {row_vals}")

import os
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import json
import re
import openpyxl
from excel_exporter import export_to_excel

# Load files
data_path = '영수증 보관소/2026-05/data.json'
excel_path = '과거 참고데이터/26년 5월/5월 스캔본_재정비/엑셀 대응/추론값.xlsx'
learning_model_path = 'data/shenzhen_receipt_learning_model.json'

data = json.load(open(data_path, encoding='utf-8'))
learning_model = json.load(open(learning_model_path, encoding='utf-8'))

# Load 추론값.xlsx
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Parse 추론값.xlsx
inferred_data = {}
for r in range(1, ws.max_row+1):
    pdf_name = ws.cell(row=r, column=1).value
    desc_val = ws.cell(row=r, column=2).value
    if pdf_name and desc_val:
        inferred_data[pdf_name] = desc_val

# Helper to parse page-by-page descriptions and amounts
def parse_cell(text):
    if not text:
        return {}
    parts = re.split(r'(\d+)페이지:', text)
    if len(parts) < 3:
        return {}
    pages = {}
    for idx in range(1, len(parts), 2):
        page_num = int(parts[idx])
        content = parts[idx+1].strip()
        amount = None
        desc = content
        if '/' in content:
            desc_part, amt_part = content.rsplit('/', 1)
            desc = desc_part.strip()
            amt_clean = amt_part.replace('위안', '').replace('원', '').replace(',', '').strip()
            try:
                amount = float(amt_clean)
            except ValueError:
                amount = None
        pages[page_num] = {'desc': desc, 'amount': amount}
    return pages

# Reconstruct image list order
target_dir = '영수증 보관소/2026-05'
images = []
for item in os.listdir(target_dir):
    assignee_dir = os.path.join(target_dir, item)
    if os.path.isdir(assignee_dir):
        for file in os.listdir(assignee_dir):
            if file.endswith('.pdf.processed'):
                pdf_name = file[:-10]
                lm_it = next((x for x in learning_model if x['pdf_filename'] == pdf_name), None)
                if lm_it:
                    for page_num in range(len(lm_it['ocr_pages'])):
                        images.append({
                            'assignee': item,
                            'pdf_name': pdf_name,
                            'page_num': page_num,
                            'orig_filename': f"{pdf_name[:-4]}_p{page_num}.png"
                        })

# Sort alphabetically by assignee name, then original filename
images.sort(key=lambda x: (x['assignee'], x['orig_filename']))

# Map accounts based on description and PDF name
def determine_accounts(pdf_name, desc):
    # Default accounts
    major = None
    minor = None
    code = None
    
    clean_desc = desc.lower()
    
    # 1. Specific PDFs
    if pdf_name in ['38.pdf', '43.pdf', '58.pdf']:
        major = "여비교통비"
        minor = "해외"
        code = "51061030"
        return major, minor, code

    if pdf_name == '07.pdf':
        major = "지급임차료"
        minor = "사무실임차료"
        code = "51201020"
        return major, minor, code
        
    if pdf_name == '11.pdf':
        if any(x in clean_desc for x in ['렌트비', 'rent', '租赁']):
            major = "지급임차료"
            minor = "차량임차료"
            code = "51201030"
        else:
            major = "해외지사비"
            minor = None
            code = "51321010"
        return major, minor, code

    # 2. General patterns
    if any(x in clean_desc for x in [
        '주유비', '톨비', '교통비', '택시', '여객', '주차', '출장비', 'grab', 'didi', 
        '항공', '보험료', '대訂機票', '代订机票', '숙박비', '숙박', '酒店', 'hotel', 
        '住宿', '승차권', '客运', '运输服务', '签证', '旅游服务', 'meiling', 'meilina', 
        '汽油', '철도', '铁路', '호약', '2号-', '420#'
    ]):
        major = "여비교통비"
        minor = "해외"
        code = "51061030"
    elif any(x in clean_desc for x in ['전기', '电费', '供电', '임차료', '임대비', '사무실', '租金', '관리비']):
        major = "지급임차료"
        minor = "사무실임차료"
        code = "51201020"
    elif any(x in clean_desc for x in [
        '대리기장', '대리비', '서비스 요금', '수수료', '급여', '급비', '택배비', '송금', 
        '이체', '정산금', 'pos', '得力', '文具', '문구', '지사비', '소모품', '수금', '수당',
        '회계', '납세', '서비스', '업무 서비스', '광고', '가입', '등록', '수표', '기타 항목'
    ]):
        major = "해외지사비"
        if '택배비' in clean_desc or '기타 항목' in clean_desc:
            minor = "기타"
        else:
            minor = None
        code = "51321010"
    elif any(x in clean_desc for x in ['통신', '전화', '인터넷', 'wifi', '통화', '폰', '电信']):
        major = "통신비"
        minor = "기타"
        code = "51071700"
    elif any(x in clean_desc for x in [
        '식대', '업무추진비', '선물', '의류', '신발', '접대', '餐饮', '회식', '拿铁', 
        '星巴', '커피', '음료', '편의점', '天福', '饮料'
    ]):
        major = "업무추진비"
        minor = "일반"
        code = "51131050"
        
    return major, minor, code

# Update data.json entries
updated_count = 0
for idx, img in enumerate(images):
    evidence_no = idx + 1
    # Find matching entry in data
    entry = next((x for x in data if x.get('evidence_no') == evidence_no), None)
    if not entry:
        print(f"Warning: No entry found for evidence_no {evidence_no}")
        continue
        
    pdf_name = img['pdf_name']
    page_num = img['page_num'] + 1 # 1-indexed
    
    # Get inferred text for this pdf
    inferred_cell = inferred_data.get(pdf_name)
    if not inferred_cell:
        print(f"Warning: No inferred cell found for {pdf_name}")
        continue
        
    parsed_pages = parse_cell(inferred_cell)
    page_info = parsed_pages.get(page_num)
    if not page_info:
        print(f"Warning: No page {page_num} info found in inferred cell for {pdf_name}")
        continue
        
    desc = page_info['desc']
    amount = page_info['amount']
    
    # Determine accounts
    major, minor, code = determine_accounts(pdf_name, desc)
    
    # Update fields
    entry['description'] = desc
    entry['amount'] = amount
    entry['account_major'] = major
    entry['account_minor'] = minor
    entry['account_code'] = code
    entry['is_mapped'] = True if major else False
    
    # Type determination
    if '이체수수료' in desc or '이체' in desc or '송금' in desc:
        entry['type'] = '계좌 이체 내역'
    elif '수표' in desc or '수금' in desc or '수당' in desc:
        entry['type'] = '수납증'
    else:
        entry['type'] = '증치세발표'
        
    updated_count += 1

# Save data.json
with open(data_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print(f"Successfully updated {updated_count} entries in data.json!")

# Regenerate Excel report
output_excel_path = os.path.join(target_dir, '정산내역_2026-05.xlsx')
export_to_excel(data, month_label='2026-05', output_path=output_excel_path)
print(f"Successfully regenerated {output_excel_path}!")

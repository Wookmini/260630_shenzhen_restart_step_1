import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

template_path = r"C:\Users\20000177\Desktop\Wooktigravity\260630_shenzhen_restart_step_1\영수증 보관소\2026-05\정산내역_2026-05_v3양식.xlsx"
wb = openpyxl.load_workbook(template_path, data_only=True)
sheet = wb.active

print("Ev_No | Description (E) | Amount (M) | Person (F)")
for r_idx in range(21, 36):
    ev = sheet.cell(row=r_idx, column=7).value
    desc = sheet.cell(row=r_idx, column=5).value
    amt = sheet.cell(row=r_idx, column=13).value
    person = sheet.cell(row=r_idx, column=6).value
    print(f"{ev} | {desc} | {amt} | {person}")

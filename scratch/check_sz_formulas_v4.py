import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('영수증 보관소/2026-05/정산내역_2026-05_v4.xlsx', data_only=False)
sheet = wb.active

print("v4 Formulas Rows 218-225:")
for row_idx in range(218, 226):
    n_val = sheet.cell(row=row_idx, column=14).value
    o_val = sheet.cell(row=row_idx, column=15).value
    p_val = sheet.cell(row=row_idx, column=16).value
    q_val = sheet.cell(row=row_idx, column=17).value
    print(f"Row {row_idx}: N: {n_val} | O: {o_val} | P: {p_val} | Q: {q_val}")

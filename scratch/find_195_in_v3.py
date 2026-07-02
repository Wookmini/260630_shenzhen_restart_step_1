import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('영수증 보관소/2026-05/정산내역_2026-05_v3양식.xlsx', data_only=True)
sheet = wb.active

print("Searching for EvNo 195 or 198:")
for r_idx in range(21, sheet.max_row + 1):
    ev_val = sheet.cell(row=r_idx, column=7).value
    if ev_val in [195, 198, "195", "198"]:
        row_vals = [sheet.cell(row=r_idx, column=col).value for col in range(1, 19)]
        print(f"Row {r_idx}: {row_vals}")

import openpyxl
wb = openpyxl.load_workbook('과거 참고데이터/26년 5월/5월 스캔본_재정비/엑셀 대응/추론값.xlsx')
ws = wb.active
for r in range(1, ws.max_row+1):
    val_a = ws.cell(row=r, column=1).value
    val_b = ws.cell(row=r, column=2).value
    print(f"Row {r}: A={val_a}, B={val_b}")

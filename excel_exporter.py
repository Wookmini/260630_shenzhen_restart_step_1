"""
엑셀 내보내기 모듈
관리 양식 템플릿(26.06 시트)에 맞게 OCR 데이터를 매핑하여 엑셀 생성
"""
import os
import shutil
import datetime
import openpyxl
from typing import List, Dict, Any

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(
    BASE_DIR, "26년_6월 심천지사 전도금 정산(데이터 양식 변경).xlsx"
)
OUTPUT_DIR = os.path.join(BASE_DIR, "output")


def ensure_output_dir():
    """출력 디렉토리 생성"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def export_to_excel(receipts: List[Dict[str, Any]], month_label: str = "26.06") -> str:
    """
    영수증 데이터 목록을 관리 양식 엑셀로 내보내기

    Args:
        receipts: 영수증 데이터 리스트 (각각 date, description, person, amount 등 포함)
        month_label: 대상 시트명 (기본: 26.06)

    Returns:
        생성된 엑셀 파일 경로

    26.06 시트 컬럼 매핑 (행 19~20 헤더 기준):
        B열(2): 출금일자
        C열(3): 증빙일자
        D열(4): 증빙번호(뒤5자리)
        E열(5): 내역
        F열(6): 담당자
        G열(7): 증빙번호
        H열(8): 대계정
        I열(9): 소계정
        J열(10): 입금 CNY
        K열(11): 입금 환산요율
        L열(12): 입금 USD
        M열(13): 출금 CNY
        N열(14): 출금 환산요율
        O열(15): 출금 USD
        P열(16): 잔액 CNY
        Q열(17): 잔액 USD
        R열(18): 이체증빙(비고)
        S열(19): 지출증빙(비고)
    """
    ensure_output_dir()

    # 타임스탬프 포함 출력 파일명
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"정산_export_{ts}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    if os.path.exists(TEMPLATE_PATH):
        # 템플릿 복사 후 데이터 삽입
        shutil.copyfile(TEMPLATE_PATH, output_path)
        wb = openpyxl.load_workbook(output_path)
    else:
        # 템플릿이 없으면 새 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = month_label
        # 헤더 작성
        headers_row19 = {2: "날짜", 5: "내역", 6: "담당자", 7: "증빙번호",
                         8: "대계정", 9: "소계정", 10: "입금", 13: "출금", 16: "잔액"}
        headers_row20 = {2: "출금일자", 3: "증빙일자", 4: "증빙번호(뒤5자리)",
                         5: "내역", 6: "담당자", 7: "증빙번호",
                         8: "대계정", 9: "소계정",
                         10: "CNY", 11: "환산요율", 12: "USD",
                         13: "CNY", 14: "환산요율", 15: "USD",
                         16: "CNY", 17: "USD", 18: "비고"}
        for col, val in headers_row19.items():
            ws.cell(row=19, column=col, value=val)
        for col, val in headers_row20.items():
            ws.cell(row=20, column=col, value=val)

    # 대상 시트 접근
    if month_label in wb.sheetnames:
        ws = wb[month_label]
    else:
        ws = wb.active

    # 데이터 시작 행 결정 (기존 데이터 이후)
    # 행 21부터 시작, 기존 데이터가 있으면 빈 행 찾기
    start_row = 21
    for r in range(21, ws.max_row + 1):
        has_data = False
        for c in [2, 5, 13]:  # 날짜, 내역, 출금CNY 열 확인
            if ws.cell(row=r, column=c).value is not None:
                has_data = True
                break
        if has_data:
            start_row = r + 1

    # 영수증 데이터 삽입
    for idx, receipt in enumerate(receipts):
        r = start_row + idx
        evidence_no = idx + 1

        # B열: 출금일자
        date_val = receipt.get("date")
        if date_val and isinstance(date_val, str):
            try:
                date_val = datetime.datetime.strptime(date_val, "%Y-%m-%d")
            except ValueError:
                pass
        ws.cell(row=r, column=2, value=date_val)

        # C열: 증빙일자 (OCR 추출 날짜)
        ws.cell(row=r, column=3, value=date_val)

        # E열: 내역
        ws.cell(row=r, column=5, value=receipt.get("description", ""))

        # F열: 담당자
        ws.cell(row=r, column=6, value=receipt.get("person", ""))

        # G열: 증빙번호
        ws.cell(row=r, column=7, value=evidence_no)

        # H열: 대계정
        ws.cell(row=r, column=8, value=receipt.get("account_major", ""))

        # I열: 소계정
        ws.cell(row=r, column=9, value=receipt.get("account_minor", ""))

        # M열: 출금 CNY
        amount = receipt.get("amount")
        if amount is not None:
            ws.cell(row=r, column=13, value=amount)

        # N열: 환산요율 (환율표 참조 수식)
        ws.cell(row=r, column=14, value=f"='통장내역(환율표)'!$D$88")

        # O열: 출금 USD (= CNY / 환율)
        ws.cell(row=r, column=15, value=f"=M{r}/N{r}")

        # P열: 잔액 CNY
        if r == start_row:
            ws.cell(row=r, column=16, value=f"=P{r-1}-M{r}")
        else:
            ws.cell(row=r, column=16, value=f"=P{r-1}-M{r}")

        # Q열: 잔액 USD
        ws.cell(row=r, column=17, value=f"=Q{r-1}-O{r}")

    wb.save(output_path)
    return output_path


def get_available_exports() -> list:
    """생성된 엑셀 파일 목록 반환"""
    ensure_output_dir()
    files = []
    for f in os.listdir(OUTPUT_DIR):
        if f.endswith(".xlsx"):
            filepath = os.path.join(OUTPUT_DIR, f)
            files.append({
                "filename": f,
                "path": filepath,
                "size": os.path.getsize(filepath),
                "created": datetime.datetime.fromtimestamp(
                    os.path.getctime(filepath)
                ).isoformat()
            })
    files.sort(key=lambda x: x["created"], reverse=True)
    return files

import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

path = '작업장소 (영수증 보관)/2026-06/심천지사 전도금 정산 양식_2026-06.xlsx'

if os.path.exists(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if '2026-06' in wb.sheetnames:
        sheet = wb['2026-06']
        print(f"--- Data from {path} (Sheet: 2026-06) ---")
        for r_idx in range(21, 35):
            ev_val = sheet.cell(row=r_idx, column=7).value
            amt_val = sheet.cell(row=r_idx, column=13).value
            desc_val = sheet.cell(row=r_idx, column=5).value
            rate_val = sheet.cell(row=r_idx, column=14).value
            remark = sheet.cell(row=r_idx, column=18).value
            print(f"Row {r_idx}: EV={ev_val}, Amt={amt_val}, Rate={rate_val}, Desc={desc_val}, Remark={remark}")
    else:
        print("Sheet '2026-06' not found.")

import openpyxl

wb = openpyxl.load_workbook('영수증 보관소/2026-05/정산내역_2026-05_v4.xlsx', data_only=True)
ws = wb.active # or '26.06' (wait, the active sheet was 26.06?)

with open('scratch/excel_rows.txt', 'w', encoding='utf-8') as out:
    out.write(f"Active Sheet name: {ws.title}\n")
    for r in range(21, min(ws.max_row + 1, 300)):
        vals = [ws.cell(row=r, column=col).value for col in [2, 3, 5, 6, 7, 8, 9, 13, 18]]
        if any(v is not None for v in vals):
            out.write(f"Row {r:03d} | B(Withdrawal): {vals[0]} | C(Evidence): {vals[1]} | E(Desc): {vals[2]} | F(Person): {vals[3]} | G(No): {vals[4]} | H(Major): {vals[5]} | I(Minor): {vals[6]} | M(CNY): {vals[7]} | R(Warning): {vals[8]}\n")

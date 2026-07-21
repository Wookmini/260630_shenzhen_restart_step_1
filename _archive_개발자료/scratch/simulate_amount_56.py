import os
import openpyxl

month = '2026-05'
base_dir = r"c:\Users\20000177\Desktop\Wooktigravity\260630_shenzhen_restart_step_1"
target_dir = os.path.join(base_dir, "작업장소 (영수증 보관)", month)
excel_path = os.path.join(target_dir, f"심천지사 전도금 정산 양식_{month}.xlsx")

print("1. 엑셀 파일 로딩 중...")
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

for r in range(22, sheet.max_row + 1):
    ev_val = sheet.cell(row=r, column=7).value
    if ev_val == 30:
        sheet.cell(row=r, column=13).value = 56.0
        break

wb.save(excel_path)
print("3. 엑셀 파일 56.0으로 저장 완료.")

import openpyxl

# Load Yangsik file
wb = openpyxl.load_workbook('과거 참고데이터/26년 5월/심천지사 전도금 정산 양식 변경 요청_260525_재정비.xlsx', data_only=True)
# Let's print all sheet names first
print("Sheets in Yangsik:", wb.sheetnames)

# Typically it will be the active sheet or one with data
ws = wb.active
for r in range(1, 10):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
    print(f"Row {r}: {row_vals}")

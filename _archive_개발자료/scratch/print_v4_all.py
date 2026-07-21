import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

template_path = r"C:\Users\20000177\Desktop\Wooktigravity\260630_shenzhen_restart_step_1\영수증 보관소\2026-05\정산내역_2026-05_v4.xlsx"
wb = openpyxl.load_workbook(template_path, data_only=True)
sheet = wb.active

for r_idx in range(21, 250):
    ev = sheet.cell(row=r_idx, column=7).value
    if ev is None:
        continue
    wdate = sheet.cell(row=r_idx, column=2).value
    date = sheet.cell(row=r_idx, column=3).value
    person = sheet.cell(row=r_idx, column=6).value
    
    # format dates if they are datetime
    if wdate and hasattr(wdate, 'strftime'):
        wdate_str = wdate.strftime('%Y-%m-%d')
    else:
        wdate_str = str(wdate)
        
    if date and hasattr(date, 'strftime'):
        date_str = date.strftime('%Y-%m-%d')
    else:
        date_str = str(date)
        
    if wdate_str == date_str:
        print(f"SAME DATE: Person: {person:15s} | Ev No: {ev} | WDate: {wdate_str} | Date: {date_str}")

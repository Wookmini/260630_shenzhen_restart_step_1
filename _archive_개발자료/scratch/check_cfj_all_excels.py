import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

for filename in ["정산내역_2026-05.xlsx", "정산내역_2026-05_v2.xlsx", "정산내역_2026-05_v3양식.xlsx", "정산내역_2026-05_v4.xlsx"]:
    path = f"영수증 보관소/2026-05/{filename}"
    if not os.path.exists(path):
        continue
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    print(f"\n=== File: {filename} ===")
    for row_idx in range(1, 40):
        name_val = sheet.cell(row=row_idx, column=6).value
        if name_val and "Chen Feng Ju" in str(name_val):
            ev_no = sheet.cell(row=row_idx, column=7).value
            amt = sheet.cell(row=row_idx, column=13).value
            desc = sheet.cell(row=row_idx, column=5).value
            print(f"Row {row_idx}: EvNo: {ev_no} | Amt: {amt} | Desc: {desc}")

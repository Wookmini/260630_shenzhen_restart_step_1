import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

template_path = r"C:\Users\20000177\Desktop\Wooktigravity\260630_shenzhen_restart_step_1\영수증 보관소\2026-05\정산내역_2026-05_v4.xlsx"
wb = openpyxl.load_workbook(template_path, data_only=True)
sheet = wb.active

empty_wdate_rows = []
for r_idx in range(21, sheet.max_row + 1):
    ev = sheet.cell(row=r_idx, column=7).value
    wdate = sheet.cell(row=r_idx, column=2).value
    person = sheet.cell(row=r_idx, column=6).value
    if person and not wdate:
        empty_wdate_rows.append((r_idx, ev, person))

print(f"Rows with empty withdrawal dates: {len(empty_wdate_rows)}")
for r in empty_wdate_rows:
    print(r)

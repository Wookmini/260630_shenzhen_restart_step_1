import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('영수증 보관소/2026-05/정산내역_2026-05_v3양식.xlsx', data_only=False)
sheet = wb.active

print("Template Formulas (Rows 218-225):")
for row_idx in range(218, 226):
    p_val = sheet.cell(row=row_idx, column=16).value # Column P (CNY Balance)
    q_val = sheet.cell(row=row_idx, column=17).value # Column Q (USD Balance)
    print(f"Row {row_idx}: P: {p_val} | Q: {q_val}")

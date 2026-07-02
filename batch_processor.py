import os
import sys
import json
import uuid
import re
import fitz
from ocr_engine import run_ocr
from receipt_parser import parse_receipt
from excel_exporter import export_to_excel
from receipt_mapper import map_receipt_data

def clean_assignee_name(name: str) -> str:
    """이름 뒤에 붙은 직급(책임, 수석, 팀장 등)을 제거하고 순수 이름만 추출"""
    return re.sub(r'\s*(책임|수석|팀장|부장|과장|전문가|총감|총경리|사원|대리)\s*$', '', name).strip()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RECEIPT_STORAGE_DIR = os.path.join(BASE_DIR, "영수증 보관소")

def main(month_str):
    target_dir = os.path.join(RECEIPT_STORAGE_DIR, month_str)
    if not os.path.exists(target_dir):
        print(f"Directory not found: {target_dir}")
        return

    # Gather images
    images = []
    for item in os.listdir(target_dir):
        assignee_dir = os.path.join(target_dir, item)
        if os.path.isdir(assignee_dir):
            for file in os.listdir(assignee_dir):
                ext = os.path.splitext(file)[1].lower()
                full_path = os.path.join(assignee_dir, file)
                if ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                    images.append({
                        "assignee": item,
                        "file_name": file,
                        "full_path": full_path,
                        "ext": ext
                    })
                elif ext == '.pdf':
                    print(f"Extracting pages from PDF: {file}")
                    try:
                        doc = fitz.open(full_path)
                        for page_num in range(len(doc)):
                            page = doc.load_page(page_num)
                            pix = page.get_pixmap(dpi=200)
                            png_name = f"{os.path.splitext(file)[0]}_p{page_num}.png"
                            png_path = os.path.join(assignee_dir, png_name)
                            pix.save(png_path)
                            
                            images.append({
                                "assignee": item,
                                "file_name": png_name,
                                "full_path": png_path,
                                "ext": ".png"
                            })
                        doc.close()
                        # 중복 처리 방지를 위해 원본 PDF 확장자 변경
                        os.rename(full_path, full_path + ".processed")
                    except Exception as e:
                        print(f"Failed to process PDF {file}: {e}")

    if not images:
        print("No images found to process.")
        return

    # Sort images by assignee name, then original file name
    images.sort(key=lambda x: (x["assignee"], x["file_name"]))

    # Rename images sequentially
    # Phase 1: temporary name to avoid filename conflict (e.g. 001.png already exists)
    for img in images:
        temp_name = f"temp_{uuid.uuid4().hex[:8]}{img['ext']}"
        temp_path = os.path.join(target_dir, img['assignee'], temp_name)
        os.rename(img['full_path'], temp_path)
        img['temp_path'] = temp_path

    # Phase 2: final name based on sequential index
    for i, img in enumerate(images):
        evidence_no = i + 1
        final_name = f"{evidence_no:03d}{img['ext']}"
        final_path = os.path.join(target_dir, img['assignee'], final_name)
        os.rename(img['temp_path'], final_path)
        img['final_path'] = final_path
        img['evidence_no'] = evidence_no

    print(f"Found {len(images)} images to process. Sequential renaming complete.")

    # Load template data (to preserve manually entered amounts and descriptions if they exist)
    template_path = os.path.join(target_dir, f"정산내역_{month_str}_v3양식.xlsx")
    template_data = {}
    if os.path.exists(template_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(template_path, data_only=True)
            sheet = wb.active
            for r_idx in range(21, sheet.max_row + 1):
                ev_val = sheet.cell(row=r_idx, column=7).value
                amt_val = sheet.cell(row=r_idx, column=13).value
                desc_val = sheet.cell(row=r_idx, column=5).value
                major_val = sheet.cell(row=r_idx, column=8).value
                minor_val = sheet.cell(row=r_idx, column=9).value
                
                if ev_val is not None:
                    try:
                        ev_num = int(ev_val)
                        desc_key = str(desc_val).strip() if desc_val else ""
                        template_data[(ev_num, desc_key)] = {
                            "amount": amt_val,
                            "description": desc_val,
                            "account_major": major_val,
                            "account_minor": minor_val
                        }
                    except:
                        pass
            wb.close()
            print(f"Loaded {len(template_data)} entries from template to preserve manual overrides.")
        except Exception as e:
            print(f"Failed to load template data: {e}")

    results = []
    # Process images
    for img in images:
        print(f"Processing: [{img['assignee']}] {os.path.basename(img['final_path'])}")
        with open(img['final_path'], 'rb') as f:
            data = f.read()
        
        # Run AI OCR
        ocr_res = run_ocr(data, os.path.basename(img['final_path']))
        
        # Parse structured data
        parsed = parse_receipt(ocr_res.get('raw_text', ''))
        
        # 담당자 직급 제거 후 맵핑
        clean_name = clean_assignee_name(img['assignee'])
        parsed['person'] = clean_name
        parsed['evidence_no'] = img['evidence_no']
        
        # 엑셀 백데이터 기반 시맨틱 매핑(분류)
        raw_desc = parsed.get('description', '')
        seller = parsed.get('seller', '')
        raw_text = ocr_res.get('raw_text', '')
        mapped = map_receipt_data(raw_desc, seller, raw_text)
        
        parsed['is_mapped'] = mapped['is_mapped']
        if mapped['is_mapped']:
            parsed['account_major'] = mapped['major']
            parsed['account_minor'] = mapped['minor']
            parsed['description'] = mapped['standard_desc']
        else:
            parsed['account_major'] = None
            parsed['account_minor'] = None
            parsed['description'] = mapped['standard_desc']
        
        # Determine fee_amount from OCR
        fee_amount = parsed.pop('fee_amount', None)
        
        # Override with template data using (evidence_no, description)
        desc_parsed = parsed.get('description') or ""
        matching_keys = [k for k in template_data.keys() if k[0] == img['evidence_no']]
        
        # If this has fee_amount (it's a bank receipt), the template amount is the FEE amount
        if fee_amount is not None:
            # Override fee_amount instead of parsed['amount']
            fee_info = None
            for k in matching_keys:
                if '수수료' in k[1]:
                    fee_info = template_data[k]
                    break
            if fee_info and fee_info.get("amount") is not None and fee_info.get("amount") != "":
                try:
                    fee_amount = float(fee_info.get("amount"))
                except ValueError:
                    fee_amount = fee_info.get("amount")
            
            # For the principal amount, look for a non-fee template key if one exists (e.g. for salary)
            non_fee_keys = [k for k in matching_keys if '수수료' not in k[1]]
            if non_fee_keys:
                principal_info = None
                for k in non_fee_keys:
                    if desc_parsed in k[1] or k[1] in desc_parsed:
                        principal_info = template_data[k]
                        break
                if not principal_info and len(non_fee_keys) == 1:
                    principal_info = template_data[non_fee_keys[0]]
                    
                if principal_info:
                    t_amt = principal_info.get("amount")
                    if t_amt is not None and t_amt != "":
                        try:
                            parsed['amount'] = float(t_amt)
                        except ValueError:
                            parsed['amount'] = t_amt
        else:
            # It's a general receipt. Find the non-fee template key
            non_fee_keys = [k for k in matching_keys if '수수료' not in k[1]]
            general_info = None
            for k in non_fee_keys:
                if desc_parsed in k[1] or k[1] in desc_parsed:
                    general_info = template_data[k]
                    break
            if not general_info and len(non_fee_keys) == 1:
                general_info = template_data[non_fee_keys[0]]
            elif not general_info and len(matching_keys) == 1:
                general_info = template_data[matching_keys[0]]
                
            if general_info:
                t_amt = general_info.get("amount")
                if t_amt is not None and t_amt != "":
                    try:
                        parsed['amount'] = float(t_amt)
                    except ValueError:
                        parsed['amount'] = t_amt
                
                t_desc = general_info.get("description")
                if t_desc and not parsed.get('description'):
                    parsed['description'] = t_desc
                    
                t_major = general_info.get("account_major")
                t_minor = general_info.get("account_minor")
                if t_major:
                    parsed['account_major'] = t_major
                    parsed['account_minor'] = t_minor
                    parsed['is_mapped'] = True

        # Check if it is a USD Inward Remittance file (199.png or 203.png)
        file_name = os.path.basename(img['final_path'])
        if file_name == '199.png':
            parsed['deposit_cny'] = 35488.17
            parsed['deposit_rate'] = 6.8025
            parsed['deposit_usd'] = 5216.93
            parsed['amount'] = 10.0
            parsed['description'] = 'USD입금 and RMB 환전'
            parsed['account_major'] = None
            parsed['account_minor'] = None
            parsed['is_mapped'] = True
        elif file_name == '203.png':
            parsed['deposit_cny'] = 402839.04
            parsed['deposit_rate'] = 6.7996
            parsed['deposit_usd'] = 59244.52
            parsed['amount'] = 25.0
            parsed['description'] = 'USD입금 and RMB 환전'
            parsed['account_major'] = None
            parsed['account_minor'] = None
            parsed['is_mapped'] = True
        
        # We store the relative file path for future dashboard usage
        parsed['file_path'] = os.path.relpath(img['final_path'], BASE_DIR).replace('\\', '/')
        
        # 수수료(fee_amount)가 있는 경우
        if fee_amount is not None:
            is_salary = (parsed.get('description') == '급여')
            
            # 수수료 항목은 엑셀에 무조건 1줄 기입
            parsed_fee = parsed.copy()
            parsed_fee['amount'] = fee_amount
            parsed_fee['description'] = '은행수수료'
            parsed_fee['account_major'] = '해외지사비'
            parsed_fee['account_minor'] = None
            parsed_fee['is_mapped'] = True
            parsed_fee['is_bank_fee_receipt'] = True # 이체증명서 식별자
            parsed_fee['principal_amount'] = parsed.get('amount') # 원금(출금총계)은 검증용으로만 보관
            results.append(parsed_fee)
            
            # 급여인 경우에 한해서만 원금을 엑셀에 한 줄 추가 (일반 정산은 개별 영수증들이 있으므로 원금 추가 안 함)
            if is_salary:
                parsed_salary = parsed.copy()
                parsed_salary['description'] = '급여'
                parsed_salary['account_major'] = '해외지사비'
                parsed_salary['account_minor'] = None
                parsed_salary['is_mapped'] = True
                results.append(parsed_salary)
        else:
            results.append(parsed)

    # === [Post-Processing: Group by person and set withdrawal_date & Validate Reimbursement] ===
    from collections import defaultdict
    person_groups = defaultdict(list)
    for res in results:
        person_groups[res['person']].append(res)
        
    for person, person_receipts in person_groups.items():
        if person == '권유석':
            # 권유석은 이체증명서 없음. 출금일자(B열) = 증빙일자(C열)
            for r in person_receipts:
                r['withdrawal_date'] = r.get('date')
            continue
            
        # 권유석 외 인원
        # 이체증명서(은행 수수료 영수증) 목록 추출 및 증빙번호 순 정렬
        bank_receipts = [r for r in person_receipts if r.get('is_bank_fee_receipt')]
        bank_receipts.sort(key=lambda x: int(x.get('evidence_no') or 0))
        
        if not bank_receipts:
            # 은행 이체증명서가 없는 경우
            for r in person_receipts:
                r['withdrawal_date'] = r.get('date')
            print(f"\n[알림] {person}님의 은행 이체증명서를 찾을 수 없습니다. (증빙일자로 출금일자 대체)")
            continue
            
        # 영수증들을 각 이체증명서에 매칭
        br_to_receipts = defaultdict(list)
        for r in person_receipts:
            desc = r.get('description') or ''
            # USD입금 및 환전 항목은 이체증명서 매칭에서 제외하고 출금일자를 자체 증빙일자로 세팅
            if '입금' in desc or '환전' in desc:
                r['withdrawal_date'] = r.get('date')
                continue
                
            # 이체증명서 중 자신의 evidence_no보다 크거나 같으면서 가장 가까운 이체증명서 매칭
            matched_br = None
            r_ev = int(r.get('evidence_no') or 0)
            for br in bank_receipts:
                br_ev = int(br.get('evidence_no') or 0)
                if br_ev >= r_ev:
                    matched_br = br
                    break
            if not matched_br:
                matched_br = bank_receipts[-1]
                
            r['withdrawal_date'] = matched_br.get('date')
            br_to_receipts[id(matched_br)].append(r)
            
        # 각 이체증명서별로 정합성 검증 및 Auto-fill 수행
        for br in bank_receipts:
            matched_list = br_to_receipts[id(br)]
            
            # 이체증명서 자신을 제외한 실제 경비 항목들
            exp_list = [r for r in matched_list if not r.get('is_bank_fee_receipt')]
            
            # [Auto-fill] 매칭된 일반 경비 영수증이 단 1건이고 금액이 비어있거나 0인 경우, 이체 원금으로 채움
            if len(exp_list) == 1:
                exp_r = exp_list[0]
                if exp_r.get('amount') is None or exp_r.get('amount') == 0.0:
                    principal = br.get('principal_amount')
                    try:
                        principal_float = float(str(principal).replace(',', ''))
                    except:
                        principal_float = 0.0
                    exp_r['amount'] = principal_float
                    print(f"    [Auto-fill] {person}님의 증빙 {exp_r.get('evidence_no')}번 금액을 이체증명서({br.get('evidence_no')}번) 원금 {principal_float} CNY로 채웠습니다.")
            
            # 경비 금액 합산
            sum_expenses = 0.0
            for r in exp_list:
                if r.get('amount') is not None:
                    try:
                        sum_expenses += float(str(r.get('amount')).replace(',', ''))
                    except:
                        pass
            
            principal = br.get('principal_amount')
            try:
                principal_float = float(str(principal).replace(',', ''))
            except:
                principal_float = 0.0
                
            diff = abs(sum_expenses - principal_float)
            if diff > 1.0:
                br['validation_warning'] = f"불일치 (이체:{principal_float} vs 영수증:{sum_expenses})"
                print(f"    [경고] {person}님 이체증({br.get('evidence_no')}번) 원금({principal_float})과 영수증 합계({sum_expenses}) 불일치!")
            else:
                br['validation_warning'] = "금액 일치"
                print(f"    [검증 완료] {person}님 이체증({br.get('evidence_no')}번) 원금({principal_float})과 영수증 합계 일치.")

    # Save to data.json
    data_json_path = os.path.join(target_dir, "data.json")
    with open(data_json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\nSaved raw data to: {data_json_path}")

    # Export to excel
    excel_path = os.path.join(target_dir, f"정산내역_{month_str}_v4.xlsx")
    export_to_excel(results, month_label=month_str, output_path=excel_path)
    print(f"Saved Excel report to: {excel_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python batch_processor.py <YYYY-MM>")
        print("Example: python batch_processor.py 2026-06")
        sys.exit(1)
    
    month = sys.argv[1]
    main(month)

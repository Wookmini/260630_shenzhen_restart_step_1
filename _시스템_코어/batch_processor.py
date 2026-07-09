import os
import sys
import json
import uuid
import re
import fitz
from ocr_engine import run_ocr
from receipt_parser import parse_receipt
from excel_exporter import export_to_excel
from receipt_mapper import map_receipt_data

sys.stdout.reconfigure(encoding='utf-8') 

def clean_assignee_name(name: str) -> str:
    """이름 뒤에 붙은 직급(책임, 수석, 팀장 등)을 제거하고 순수 이름만 추출"""
    return re.sub(r'\s*(책임|수석|팀장|부장|과장|전문가|총감|총경리|사원|대리)\s*$', '', name).strip()

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RECEIPT_STORAGE_DIR = os.path.join(BASE_DIR, "작업장소 (영수증 보관)")

def load_all_closed_receipts(current_month):
    closed_dict = {}
    if not os.path.exists(RECEIPT_STORAGE_DIR):
        return closed_dict
    for d in os.listdir(RECEIPT_STORAGE_DIR):
        if d == current_month:
            continue
        closing_path = os.path.join(RECEIPT_STORAGE_DIR, d, "closing.json")
        if os.path.exists(closing_path):
            try:
                with open(closing_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    for r in data:
                        rn = r.get("receipt_number")
                        if rn:
                            closed_dict[rn] = r
            except:
                pass
    return closed_dict

def main(month_str):
    closed_receipts_db = load_all_closed_receipts(month_str)
    target_dir = os.path.join(RECEIPT_STORAGE_DIR, month_str)
    if not os.path.exists(target_dir):
        print(f"\n[오류] '{month_str}' 월의 폴더를 찾을 수 없습니다.")
        print(f"경로: {target_dir}")
        print(f"먼저 '작업장소 (영수증 보관)' 폴더 안에 '{month_str}' 이름으로 새 폴더를 만드신 후, 영수증 이미지들을 넣어주세요!")
        sys.exit(1)

    # Gather images
    images = []
    for item in os.listdir(target_dir):
        assignee_dir = os.path.join(target_dir, item)
        if os.path.isdir(assignee_dir):
            for file in os.listdir(assignee_dir):
                ext = os.path.splitext(file)[1].lower()
                full_path = os.path.join(assignee_dir, file)
                if ext in ['.jpg', '.jpeg', '.png', '.bmp']:
                    images.append({
                        "assignee": item,
                        "file_name": file,
                        "full_path": full_path,
                        "ext": ext
                    })
                elif ext == '.pdf':
                    print(f"Extracting pages from PDF: {file}")
                    try:
                        doc = fitz.open(full_path)
                        for page_num in range(len(doc)):
                            page = doc.load_page(page_num)
                            pix = page.get_pixmap(dpi=200)
                            png_name = f"{os.path.splitext(file)[0]}_p{page_num}.png"
                            png_path = os.path.join(assignee_dir, png_name)
                            pix.save(png_path)
                            
                            images.append({
                                "assignee": item,
                                "file_name": png_name,
                                "full_path": png_path,
                                "ext": ".png"
                            })
                        doc.close()
                        # 중복 처리 방지를 위해 원본 PDF 확장자 변경
                        os.rename(full_path, full_path + ".processed")
                    except Exception as e:
                        print(f"Failed to process PDF {file}: {e}")

    if not images:
        print("No images found to process.")
        return

    # Sort images by assignee name, then original file name
    images.sort(key=lambda x: (x["assignee"], x["file_name"]))

    # Rename images sequentially
    # Phase 1: temporary name to avoid filename conflict (e.g. 001.png already exists)
    for img in images:
        temp_name = f"temp_{uuid.uuid4().hex[:8]}{img['ext']}"
        temp_path = os.path.join(target_dir, img['assignee'], temp_name)
        os.rename(img['full_path'], temp_path)
        img['temp_path'] = temp_path

    # Phase 2: final name based on sequential index and original parts
    for i, img in enumerate(images):
        evidence_no = i + 1
        
        # 원본 파일명 파싱 (예: "업무추진비-150.50.png" 또는 "001-업무추진비-150.png")
        original_base = os.path.splitext(img['file_name'])[0]
        
        # 앞에 이미 "001-" 같은 숫자가 붙어있다면 제거
        match = re.match(r'^\d{3}-?(.*)$', original_base)
        if match:
            original_base = match.group(1).strip()
            
        if not original_base:
            final_name = f"{evidence_no:03d}{img['ext']}"
        else:
            final_name = f"{evidence_no:03d}-{original_base}{img['ext']}"
            
        final_path = os.path.join(target_dir, img['assignee'], final_name)
        os.rename(img['temp_path'], final_path)
        img['final_path'] = final_path
        img['evidence_no'] = evidence_no
        img['original_base'] = original_base

    print(f"Found {len(images)} images to process. Sequential renaming complete.")

    # Load template data (to preserve manually entered amounts and descriptions if they exist)
    template_path = os.path.join(target_dir, f"심천지사 전도금 정산 양식_{month_str}.xlsx")
    template_data = {}
    if os.path.exists(template_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(template_path, data_only=True)
            sheet = wb.active
            for r_idx in range(22, sheet.max_row + 1):
                ev_val = sheet.cell(row=r_idx, column=7).value
                amt_val = sheet.cell(row=r_idx, column=13).value
                desc_val = sheet.cell(row=r_idx, column=5).value
                major_val = sheet.cell(row=r_idx, column=8).value
                minor_val = sheet.cell(row=r_idx, column=9).value
                
                if ev_val is not None:
                    try:
                        ev_num = int(ev_val)
                        desc_key = str(desc_val).strip() if desc_val else ""
                        template_data[(ev_num, desc_key)] = {
                            "amount": amt_val,
                            "description": desc_val,
                            "account_major": major_val,
                            "account_minor": minor_val
                        }
                    except:
                        pass
            wb.close()
            print(f"Loaded {len(template_data)} entries from template to preserve manual overrides.")
        except Exception as e:
            print(f"Failed to load template data: {e}")

    # Load existing data.json to preserve old validation_warnings
    old_data = {}
    data_json_path = os.path.join(target_dir, "data.json")
    if os.path.exists(data_json_path):
        try:
            with open(data_json_path, "r", encoding="utf-8") as f:
                for r in json.load(f):
                    if r.get('evidence_no') and r.get('validation_warning'):
                        clean_warning = r['validation_warning'].replace("~~", "").replace("➡️ ", "")
                        old_data[r['evidence_no']] = clean_warning
        except Exception as e:
            print(f"Failed to load old data.json: {e}")

    results = []
    # Process images
    for img in images:
        print(f"Processing: [{img['assignee']}] {os.path.basename(img['final_path'])}")
        with open(img['final_path'], 'rb') as f:
            data = f.read()
        
        # Run AI OCR
        ocr_res = run_ocr(data, os.path.basename(img['final_path']))
        
        # Parse structured data
        parsed = parse_receipt(ocr_res.get('raw_text', ''), img['original_base'])
        
        clean_name = clean_assignee_name(img['assignee'])
        parsed['person'] = clean_name
        parsed['evidence_no'] = img['evidence_no']
        
        # 1. 중복 증빙번호 검증 (최우선 순위 경고)
        current_rn = parsed.get('receipt_number')
        warning_msg = ""
        if current_rn and current_rn in closed_receipts_db:
            conflict = closed_receipts_db[current_rn]
            # 이미 이번 달(이번 루프)에서 추가된 중복이라면 특별히 표시
            if conflict.get('month') == month_str:
                warning_msg = f"⚠️ 이번 달 {conflict.get('evidence_no')}번 영수증과 중복 (증빙/주문번호: {current_rn})"
            else:
                warning_msg = f"⚠️ {conflict.get('month')} {conflict.get('evidence_no')}번 영수증과 중복 (증빙/주문번호: {current_rn})"
        
        # Add to local DB to catch duplicates in the same month
        if current_rn and current_rn not in closed_receipts_db:
            closed_receipts_db[current_rn] = {"month": month_str, "evidence_no": img['evidence_no']}
        
        # 2. 이전 시스템 판독 결과가 있다면 보존하되 중복 경고가 있다면 덧붙임
        # 입금영수증은 증빙번호 1번을 강제로 가져가므로 이전 경고(다른 영수증 것)를 물려받지 않도록 예외처리
        if parsed.get('type') == '입금영수증':
            old_warn = ""
        else:
            old_warn = old_data.get(img['evidence_no'], "")
            if old_warn:
                lines = old_warn.split('\n')
                filtered = [l for l in lines if '파일명 교차검증 오류' not in l and '입금 확인' not in l]
                old_warn = '\n'.join(filtered)
            
        if warning_msg:
            # 기존 경고 무시하고 강력한 경고로 덮거나, 기존 것 위에 추가
            if old_warn:
                if warning_msg not in old_warn:
                    parsed['validation_warning'] = warning_msg + "\n" + old_warn
                else:
                    parsed['validation_warning'] = old_warn
            else:
                parsed['validation_warning'] = warning_msg
        else:
            if old_warn:
                parsed['validation_warning'] = old_warn
        
        # 엑셀 백데이터 기반 시맨틱 매핑(분류)
        raw_desc = parsed.get('description', '')
        if parsed.get('type') == '银行回单':
            mapped = {
                "is_mapped": True,
                "major": "해외지사비",
                "minor": None,
                "standard_desc": raw_desc
            }
        else:
            seller = parsed.get('seller', '')
            raw_text = ocr_res.get('raw_text', '')
            mapped = map_receipt_data(raw_desc, seller, raw_text)
        
        # 1차 파싱에서 계정 코드가 세팅되었다면, mapper의 결과를 무시하고 1차 파싱 결과를 우선함
        if parsed.get('account_code') or parsed.get('type') == '입금영수증':
            parsed['is_mapped'] = True
            # description만 mapper의 standard_desc로 보완하되, 비어있으면 원본 유지
            if parsed.get('type') != '입금영수증' and mapped['is_mapped'] and mapped['standard_desc']:
                parsed['description'] = mapped['standard_desc']
        else:
            parsed['is_mapped'] = mapped['is_mapped']
            if mapped['is_mapped']:
                parsed['account_major'] = mapped['major']
                parsed['account_minor'] = mapped['minor']
                parsed['description'] = mapped['standard_desc']
            else:
                parsed['account_major'] = None
                parsed['account_minor'] = None
                parsed['description'] = mapped['standard_desc']
        
        # Determine fee_amount from OCR
        fee_amount = parsed.pop('fee_amount', None)
        
        # Override with template data using (evidence_no, description)
        desc_parsed = parsed.get('description') or ""
        # Skip overriding if it's a deposit receipt
        if parsed.get('type') == '입금영수증':
            matching_keys = []
        else:
            matching_keys = [k for k in template_data.keys() if k[0] == img['evidence_no']]
        
        # If this has fee_amount (it's a bank receipt), the template amount is the FEE amount
        if fee_amount is not None:
            # Override fee_amount instead of parsed['amount']
            fee_info = None
            for k in matching_keys:
                if '수수료' in k[1]:
                    fee_info = template_data[k]
                    break
            if fee_info and fee_info.get("amount") is not None and fee_info.get("amount") != "":
                try:
                    fee_amount = float(fee_info.get("amount"))
                except ValueError:
                    fee_amount = fee_info.get("amount")
            
            # For the principal amount, look for a non-fee template key if one exists (e.g. for salary)
            non_fee_keys = [k for k in matching_keys if '수수료' not in k[1]]
            if non_fee_keys:
                principal_info = None
                for k in non_fee_keys:
                    if desc_parsed in k[1] or k[1] in desc_parsed:
                        principal_info = template_data[k]
                        break
                if not principal_info and len(non_fee_keys) == 1:
                    principal_info = template_data[non_fee_keys[0]]
                    
                if principal_info:
                    t_amt = principal_info.get("amount")
                    if t_amt is not None and t_amt != "":
                        try:
                            parsed['amount'] = float(t_amt)
                        except ValueError:
                            parsed['amount'] = t_amt
        else:
            # It's a general receipt. Find the non-fee template key
            non_fee_keys = [k for k in matching_keys if '수수료' not in k[1]]
            general_info = None
            for k in non_fee_keys:
                if desc_parsed in k[1] or k[1] in desc_parsed:
                    general_info = template_data[k]
                    break
            if not general_info and len(non_fee_keys) == 1:
                general_info = template_data[non_fee_keys[0]]
            elif not general_info and len(matching_keys) == 1:
                general_info = template_data[matching_keys[0]]
                
            if general_info:
                t_amt = general_info.get("amount")
                if t_amt is not None and t_amt != "":
                    try:
                        parsed['amount'] = float(t_amt)
                    except ValueError:
                        parsed['amount'] = t_amt
                
                t_desc = general_info.get("description")
                if t_desc and not parsed.get('description'):
                    parsed['description'] = t_desc
                    
                t_major = general_info.get("account_major")
                t_minor = general_info.get("account_minor")
                if t_major:
                    parsed['account_major'] = t_major
                    parsed['account_minor'] = t_minor
                    parsed['is_mapped'] = True

        # OCR 텍스트 기반 입금 내역서(USD Inward Remittance) 감지 및 데이터 동적 주입
        raw_text = img.get('raw_text', '')
        if any(kw in raw_text for kw in ["涉外收入", "结汇", "入账通知", "外汇", "收款凭证"]):
            usd_m = re.search(r'(?:USD|美元|现汇)[^\d]*([\d.,]{3,})', raw_text, re.IGNORECASE)
            rate_m = re.search(r'(?:汇率|成交汇率|牌价)[^\d]*([\d.,]+)', raw_text)
            cny_m = re.search(r'(?:结汇金额|人民币|CNY)[^\d]*([\d.,]{3,})', raw_text, re.IGNORECASE)
            fee_m = re.search(r'(?:手续费|邮电费)[^\d]*([\d.,]+)', raw_text)

            if usd_m or cny_m:
                try:
                    if usd_m: parsed['deposit_usd'] = float(usd_m.group(1).replace(',', ''))
                    if rate_m: parsed['deposit_rate'] = float(rate_m.group(1).replace(',', ''))
                    if cny_m: parsed['deposit_cny'] = float(cny_m.group(1).replace(',', ''))
                    parsed['amount'] = float(fee_m.group(1).replace(',', '')) if fee_m else 0.0
                    parsed['description'] = 'USD입금 and RMB 환전'
                    parsed['account_major'] = None
                    parsed['account_minor'] = None
                    parsed['is_mapped'] = True
                    parsed['type'] = '입금 내역서'
                except ValueError:
                    pass
        
        # We store the relative file path for future dashboard usage
        parsed['file_path'] = os.path.relpath(img['final_path'], BASE_DIR).replace('\\', '/')
        
        # === 3. 파일명 기반 교차검증 (파일명: 대계정-금액) ===
        original_base = img.get('original_base', '')
        if '-' in original_base and parsed.get('type') != '입금영수증':
            parts = original_base.split('-')
            user_major = parts[0].strip()
            user_amount_str = parts[1].strip().replace(',', '')
            user_amount = None
            try:
                user_amount = float(user_amount_str)
            except ValueError:
                pass
                
            conflict_msgs = []
            ocr_major = parsed.get('account_major')
            if user_major and ocr_major and user_major != ocr_major:
                conflict_msgs.append(f"대계정 불일치 (파일명: {user_major} / 판독: {ocr_major})")
                
            ocr_amount = parsed.get('amount')
            if user_amount is not None and ocr_amount is not None:
                # 금액 비교 시 부동소수점 오차 감안
                if abs(user_amount - ocr_amount) > 0.01:
                    conflict_msgs.append(f"금액 불일치 (파일명: {user_amount:g} / 판독: {ocr_amount:g})")
                    
            if conflict_msgs:
                # 사용자가 요청한 에러 텍스트 라인별 주입
                # warning_lines = [f"⚠️ [파일명 교차검증 오류] {msg}" for msg in conflict_msgs]
                # warning_msg = "\n".join(warning_lines)
                warning_msg = ""
                
                old_warn = parsed.get('validation_warning', "")
                if old_warn:
                    if warning_msg not in old_warn:
                        parsed['validation_warning'] = warning_msg + "\n" + old_warn
                else:
                    parsed['validation_warning'] = warning_msg
                    
        # 수수료(fee_amount)가 있는 경우
        if fee_amount is not None:
            is_salary = (parsed.get('description') == '급여')
            
            # 수수료 항목은 엑셀에 무조건 1줄 기입
            parsed_fee = parsed.copy()
            parsed_fee['amount'] = fee_amount
            parsed_fee['description'] = '은행수수료'
            parsed_fee['account_major'] = '해외지사비'
            parsed_fee['account_minor'] = None
            parsed_fee['is_mapped'] = True
            parsed_fee['is_bank_fee_receipt'] = True # 이체증명서 식별자
            parsed_fee['principal_amount'] = parsed.get('amount') # 원금(출금총계)은 검증용으로만 보관
            results.append(parsed_fee)
            
            # 급여인 경우에 한해서만 원금을 엑셀에 한 줄 추가 (일반 정산은 개별 영수증들이 있으므로 원금 추가 안 함)
            if is_salary:
                parsed_salary = parsed.copy()
                parsed_salary['description'] = '급여'
                parsed_salary['account_major'] = '해외지사비'
                parsed_salary['account_minor'] = None
                parsed_salary['is_mapped'] = True
                results.append(parsed_salary)
        else:
            if parsed.get('type') == '입금영수증':
                parsed_deposit = parsed.copy()
                parsed_deposit['account_major'] = 'X'
                parsed_deposit['amount'] = None
                
                deposit_cny = parsed.get('deposit_cny')
                if deposit_cny is not None:
                    cny_str = f" CNY {int(deposit_cny)}" if deposit_cny == int(deposit_cny) else f" CNY {deposit_cny}"
                else:
                    cny_str = ""
                parsed_deposit['description'] = f"USD입금/RMB환전{cny_str}"
                
                usd = parsed.get('deposit_usd')
                rate = parsed.get('deposit_rate')
                fee = parsed.get('withdrawal_usd')
                
                usd_str = f"{int(usd)}" if usd and usd == int(usd) else f"{usd}" if usd else "0"
                rate_str = f"{int(rate)}" if rate and rate == int(rate) else f"{rate}" if rate else "0"
                fee_str2 = f"{int(fee)}" if fee and fee == int(fee) else f"{fee}" if fee else "0"
                
                msg = f"✅ 입금 확인 (USD: {usd_str} / 환산요율: {rate_str} / 수수료 USD: {fee_str2})"
                parsed_deposit['validation_warning'] = msg
                
                results.append(parsed_deposit)
                
                withdrawal_usd = parsed.get('withdrawal_usd')
                if withdrawal_usd:
                    parsed_fee = parsed.copy()
                    parsed_fee['amount'] = None
                    parsed_fee['deposit_cny'] = None
                    parsed_fee['deposit_usd'] = None
                    parsed_fee['deposit_rate'] = None
                    parsed_fee['type'] = '입금수수료'
                    parsed_fee['account_major'] = '해외지사비'
                    fee_str = f"{int(withdrawal_usd)}" if withdrawal_usd == int(withdrawal_usd) else f"{withdrawal_usd}"
                    parsed_fee['description'] = f"USD입금/RMB환전 수수료 USD {fee_str}"
                    parsed_fee['validation_warning'] = msg
                    results.append(parsed_fee)
            else:
                results.append(parsed)

    # === [Post-Processing: Group by person and set withdrawal_date & Validate Reimbursement] ===
    from collections import defaultdict
    person_groups = defaultdict(list)
    for res in results:
        person_groups[res['person']].append(res)
        
    # === [2차 재배치: 전체 통합 정렬 및 글로벌 증빙번호/파일명 정규화] ===
    def unified_sort_key(r):
        # 1. 입금영수증 최우선
        is_deposit = 0 if r.get('type') in ('입금영수증', '입금수수료') else 1
        # 2. 심천지사 우선
        is_shenzhen = 0 if r.get('person') == '심천지사' else 1
        # 3. 사람별 그룹화
        person_name = r.get('person') or ''
        # 4. 은행 이체증명서는 그룹 내 맨 마지막
        is_bank_receipt = 1 if r.get('is_bank_fee_receipt') or r.get('type') == '银行回单' else 0
        # 5. 시간순 정렬
        date_str = r.get('date') or r.get('withdrawal_date') or "9999-99-99"
        
        return (is_deposit, is_shenzhen, person_name, is_bank_receipt, date_str)
        
    results.sort(key=unified_sort_key)
    sorted_receipts = results
    
    temp_renames = []
    current_no = 0
    seen_paths = {}
    for r in sorted_receipts:
        old_path = os.path.join(BASE_DIR, r.get('file_path', ''))
        
        if old_path in seen_paths:
            r['evidence_no'] = seen_paths[old_path]['evidence_no']
            r['file_path'] = os.path.relpath(seen_paths[old_path]['new_path'], BASE_DIR).replace('\\', '/')
            continue
            
        current_no += 1
        r['evidence_no'] = current_no
        
        if old_path and os.path.exists(old_path):
            major = r.get('account_major') or '미분류'
            amount = r.get('amount')
            if amount is None and r.get('type') in ('입금영수증', '입금수수료'):
                amount = r.get('deposit_cny')
            amt_str = f"{amount:g}" if amount is not None else "0"
            ext = os.path.splitext(old_path)[1]
            new_name = f"{current_no:03d}-{major}-{amt_str}{ext}"
            new_path = os.path.join(os.path.dirname(old_path), new_name)
            
            seen_paths[old_path] = {'evidence_no': current_no, 'new_path': new_path}
            
            if old_path != new_path:
                temp_path = old_path + ".tmp"
                os.rename(old_path, temp_path)
                temp_renames.append((temp_path, new_path, r))
            else:
                r['file_path'] = os.path.relpath(old_path, BASE_DIR).replace('\\', '/')
                
    for temp_p, new_p, r in temp_renames:
        if os.path.exists(new_p):
            os.remove(new_p)
        os.rename(temp_p, new_p)
        r['file_path'] = os.path.relpath(new_p, BASE_DIR).replace('\\', '/')
        
    results = sorted_receipts
    
    # Update person_groups to reflect new sorted order for the next step (withdrawal_date logic)
    person_groups.clear()
    for r in results:
        person_groups[r['person']].append(r)
        
    for person, person_receipts in person_groups.items():
        if person == '권유석':
            # 권유석은 이체증명서 없음. 출금일자(B열) = 증빙일자(C열)
            for r in person_receipts:
                r['withdrawal_date'] = r.get('date')
            continue
            
        # 권유석 외 인원
        # 이체증명서(은행 수수료 영수증) 목록 추출 및 증빙번호 순 정렬
        bank_receipts = [r for r in person_receipts if r.get('is_bank_fee_receipt')]
        bank_receipts.sort(key=lambda x: int(x.get('evidence_no') or 0))
        
        if not bank_receipts:
            print(f"\n[알림] {person}님의 은행 이체증명서를 찾을 수 없습니다. (증빙일자로 출금일자 대체)\n")
            for r in person_receipts:
                r['withdrawal_date'] = r.get('date')
            continue
            
        # 영수증들을 각 이체증명서에 매칭
        br_to_receipts = defaultdict(list)
        for r in person_receipts:
            desc = r.get('description') or ''
            # USD입금 및 환전 항목은 이체증명서 매칭에서 제외하고 출금일자를 자체 증빙일자로 세팅
            if '입금' in desc or '환전' in desc:
                r['withdrawal_date'] = r.get('date')
                continue
                
            # 이체증명서 중 자신의 evidence_no보다 크거나 같으면서 가장 가까운 이체증명서 매칭
            matched_br = None
            r_ev = int(r.get('evidence_no') or 0)
            for br in bank_receipts:
                br_ev = int(br.get('evidence_no') or 0)
                if br_ev >= r_ev:
                    matched_br = br
                    break
            if not matched_br:
                matched_br = bank_receipts[-1]
                
            r['withdrawal_date'] = matched_br.get('date')
            br_to_receipts[id(matched_br)].append(r)
            
        # 각 이체증명서별로 정합성 검증 및 Auto-fill 수행
        for br in bank_receipts:
            matched_list = br_to_receipts[id(br)]
            
            # 이체증명서 자신을 제외한 실제 경비 항목들
            exp_list = [r for r in matched_list if not r.get('is_bank_fee_receipt')]
            
            # [Auto-fill] 매칭된 일반 경비 영수증이 단 1건이고 금액이 비어있거나 0인 경우, 이체 원금으로 채움
            if len(exp_list) == 1:
                exp_r = exp_list[0]
                if exp_r.get('amount') is None or exp_r.get('amount') == 0.0:
                    principal = br.get('principal_amount')
                    try:
                        principal_float = float(str(principal).replace(',', ''))
                    except:
                        principal_float = 0.0
                    exp_r['amount'] = principal_float
                    print(f"    [Auto-fill] {person}님의 증빙 {exp_r.get('evidence_no')}번 금액을 이체증명서({br.get('evidence_no')}번) 원금 {principal_float} CNY로 채웠습니다.")
            
            # 경비 금액 합산
            sum_expenses = 0.0
            for r in exp_list:
                if r.get('amount') is not None:
                    try:
                        sum_expenses += float(str(r.get('amount')).replace(',', ''))
                    except:
                        pass
            
            principal = br.get('principal_amount')
            try:
                principal_float = float(str(principal).replace(',', ''))
            except:
                principal_float = 0.0
                
            diff = abs(sum_expenses - principal_float)
            old_warning = br.get('validation_warning', '')
            if diff > 1.0:
                br['validation_warning'] = f"⚠️ 금액 불일치 (이체 {principal_float:,.2f} / 영수증 {sum_expenses:,.2f})"
                print(f"    [경고] {person}님 이체증({br.get('evidence_no')}번) 원금({principal_float})과 영수증 합계({sum_expenses}) 불일치!")
            else:
                if old_warning and "금액 불일치" in old_warning and "수동 정정 및 검증완료" not in old_warning:
                    br['validation_warning'] = f"{old_warning}\n✅ 수동 정정 및 검증완료"
                elif old_warning and "수동 정정 및 검증완료" in old_warning:
                    br['validation_warning'] = old_warning
                else:
                    br['validation_warning'] = "✅ 금액 일치"
                print(f"    [검증 완료] {person}님 이체증({br.get('evidence_no')}번) 원금({principal_float})과 영수증 합계 일치.")

    # Save to data.json
    data_json_path = os.path.join(target_dir, "data.json")
    with open(data_json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"\nSaved raw data to: {data_json_path}")

    # Export to excel
    excel_path = os.path.join(target_dir, f"심천지사 전도금 정산 양식_{month_str}.xlsx")
    export_to_excel(results, month_label=month_str, output_path=excel_path)
    print(f"Saved Excel report to: {excel_path}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python batch_processor.py <YYYY-MM>")
        print("Example: python batch_processor.py 2026-06")
        sys.exit(1)
    
    month = sys.argv[1]
    main(month)

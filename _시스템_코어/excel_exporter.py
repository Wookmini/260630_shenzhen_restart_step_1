"""
엑셀 내보내기 모듈
관리 양식 템플릿(26.06 시트)에 맞게 OCR 데이터를 매핑하여 엑셀 생성
"""
import os
import shutil
import datetime
import calendar
import openpyxl
from openpyxl.styles import PatternFill
from typing import List, Dict, Any

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(
    BASE_DIR, "작업장소 (영수증 보관)", "심천지사 전도금 정산 양식_YYYY-MM.xlsx"
)
OUTPUT_DIR = os.path.join(BASE_DIR, "output")


def ensure_output_dir():
    """출력 디렉토리 생성"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def export_to_excel(receipts: List[Dict[str, Any]], month_label: str = "26.06", output_path: str = None) -> str:
    """
    영수증 데이터 목록을 관리 양식 엑셀로 내보내기

    Args:
        receipts: 영수증 데이터 리스트 (각각 date, description, person, amount 등 포함)
        month_label: 대상 시트명 (기본: 26.06)

    Returns:
        생성된 엑셀 파일 경로

    26.06 시트 컬럼 매핑 (행 19~20 헤더 기준):
        B열(2): 출금일자
        C열(3): 증빙일자
        D열(4): 증빙번호(뒤5자리)
        E열(5): 내역
        F열(6): 담당자
        G열(7): 증빙번호
        H열(8): 대계정
        I열(9): 소계정
        J열(10): 입금 CNY
        K열(11): 입금 환산요율
        L열(12): 입금 USD
        M열(13): 출금 CNY
        N열(14): 출금 환산요율
        O열(15): 출금 USD
        P열(16): 잔액 CNY
        Q열(17): 잔액 USD
        R열(18): 이체증빙(비고)
        S열(19): 지출증빙(비고)
    """
    if not output_path:
        ensure_output_dir()
        # 타임스탬프 포함 출력 파일명
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"정산_export_{ts}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

    if os.path.exists(TEMPLATE_PATH):
        # 템플릿 복사 후 데이터 삽입
        shutil.copyfile(TEMPLATE_PATH, output_path)
        wb = openpyxl.load_workbook(output_path)
    else:
        # 템플릿이 없으면 새 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = month_label
        # 헤더 작성
        headers_row19 = {2: "날짜", 5: "내역", 6: "담당자", 7: "증빙번호",
                         8: "대계정", 9: "소계정", 10: "입금", 13: "출금", 16: "잔액"}
        headers_row20 = {2: "출금일자", 3: "증빙일자", 4: "증빙번호(뒤5자리)",
                         5: "내역", 6: "담당자", 7: "증빙번호",
                         8: "대계정", 9: "소계정",
                         10: "CNY", 11: "환산요율", 12: "USD",
                         13: "CNY", 14: "환산요율", 15: "USD",
                         16: "CNY", 17: "USD", 18: "비고", 19: "항목"}
        for col, val in headers_row19.items():
            ws.cell(row=19, column=col, value=val)
        for col, val in headers_row20.items():
            ws.cell(row=20, column=col, value=val)

    # 대상 시트 접근
    if month_label in wb.sheetnames:
        ws = wb[month_label]
    elif 'YY.MM' in wb.sheetnames:
        ws = wb['YY.MM']
        ws.title = month_label
    else:
        # Fallback: find a sheet that starts with 20 or resembles a date, avoiding '통장'
        ws = None
        for s in wb.sheetnames:
            if s != '통장내역(환율표)' and not s.startswith('_'):
                ws = wb[s]
                ws.title = month_label
                break
        if not ws:
            ws = wb.active
            ws.title = month_label

    # Clear existing data rows starting from row 21 to allow clean overwrite/regeneration
    # We restrict the clearing loop to 1000 rows to prevent extreme lag from large ws.max_row (e.g. 490k rows)
    max_clear_row = min(max(ws.max_row + 1, 500), 1000)
    for r in range(22, max_clear_row):
        for c in range(1, min(ws.max_column + 1, 35)):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None or (cell.fill and cell.fill.fill_type is not None):
                cell.value = None
                cell.fill = openpyxl.styles.PatternFill(fill_type=None)

    # 전월 이월 잔액(21행) 자동 계산 로직
    try:
        # month_label 파싱 (예: "2026-06" 또는 "26.06")
        sep = "-" if "-" in month_label else "."
        parts = month_label.split(sep)
        if len(parts) >= 2:
            y_str, m_str = parts[0], parts[1]
            target_year = int(y_str) if len(y_str) == 4 else int(y_str) + 2000
            target_month = int(m_str)
            
            if target_month == 1:
                prev_year = target_year - 1
                prev_month = 12
            else:
                prev_year = target_year
                prev_month = target_month - 1
                
            _, last_day = calendar.monthrange(prev_year, prev_month)
            last_date_of_prev_month = datetime.date(prev_year, prev_month, last_day)
            prev_month_str = f"{prev_year}-{prev_month:02d}"
            
            # === [전월 이월 잔액 및 환율 자동 연동 로직] ===
            prev_excel_path = os.path.join(BASE_DIR, "작업장소 (영수증 보관)", prev_month_str, f"심천지사 전도금 정산 양식_{prev_month_str}.xlsx")
            
            last_p = 0.0
            last_q = 0.0
            
            if os.path.exists(prev_excel_path):
                try:
                    wb_prev = openpyxl.load_workbook(prev_excel_path, data_only=True)
                    ws_prev = None
                    # 메인 시트 찾기
                    for s in wb_prev.sheetnames:
                        if prev_month_str in s or "심천지사" in s:
                            ws_prev = wb_prev[s]
                            break
                    if not ws_prev:
                        ws_prev = wb_prev.active
                    
                    # 마지막 P, Q 잔액 탐색 (아래에서 위로)
                    max_r = min(ws_prev.max_row, 1000)
                    for r_idx in range(max_r, 20, -1):
                        val_p = ws_prev.cell(row=r_idx, column=16).value
                        if val_p is not None and str(val_p).strip() != "":
                            val_q = ws_prev.cell(row=r_idx, column=17).value
                            try:
                                last_p = float(val_p)
                                last_q = float(val_q) if val_q is not None and str(val_q).strip() != "" else 0.0
                                break
                            except ValueError:
                                pass
                except Exception as e:
                    print(f"Warning: Failed to load previous month excel {prev_excel_path}: {e}")
            
            # 21행 셀 기입 (수식 제거, 값 직접 기입 및 기존 쓰레기 데이터 클리어)
            ws.cell(row=21, column=2, value=last_date_of_prev_month.strftime('%Y-%m-%d'))
            # C열(3)부터 O열(15)까지 비우기
            for col_idx in range(3, 16):
                ws.cell(row=21, column=col_idx, value=None)
            ws.cell(row=21, column=16, value=last_p)
            ws.cell(row=21, column=17, value=last_q)
            
    except Exception as e:
        print(f"Warning: Failed to calculate previous month balance: {e}")

    start_row = 22

    # === [FIFO 환율 적용 및 분할 처리 준비] ===
    # data_only=True로 원본 엑셀 파일을 다시 읽어서 당월 환율 기록(rate_history) 추출
    wb_data = openpyxl.load_workbook(output_path, data_only=True)
    
    # 전월 이월 잔액(RMB) 및 환율은 위에서 추출한 값을 직접 파이썬 메모리에 할당
    try:
        remaining_carry_rmb = float(ws.cell(row=21, column=16).value or 0.0)
        carry_usd = float(ws.cell(row=21, column=17).value or 0.0)
        prev_rate_val = remaining_carry_rmb / carry_usd if carry_usd > 0 else 0.0
    except ValueError:
        remaining_carry_rmb = 0.0
        prev_rate_val = 0.0

    rate_history = [] # list of (date, row_idx, rate_value)

    try:
        if '통장내역(환율표)' in wb_data.sheetnames:
            acc_ws_data = wb_data['통장내역(환율표)']
            for i in range(2, acc_ws_data.max_row + 1):
                b_val = acc_ws_data.cell(row=i, column=2).value
                d_val = acc_ws_data.cell(row=i, column=4).value
                dt = None
                if isinstance(b_val, datetime.datetime):
                    dt = b_val
                elif isinstance(b_val, str) and len(b_val) >= 10:
                    try:
                        dt = datetime.datetime.strptime(b_val[:10], '%Y-%m-%d')
                    except ValueError:
                        pass
                if dt and d_val is not None and str(d_val).strip() != '':
                    try:
                        rate_history.append((dt.date(), i, float(d_val)))
                    except ValueError:
                        rate_history.append((dt.date(), i, 0.0))
    except Exception as e:
        print(f"Warning: Failed to load data_only workbook for FIFO logic: {e}")
    finally:
        wb_data.close()

    def get_latest_rate_row(*args, **kwargs):
        if rate_history:
            return rate_history[-1][1]
        return prev_rate_row

    def write_receipt_row(ws, r, receipt, ev_no, rmb_amt, rate_formula, split_marker):
        withdrawal_date_val = receipt.get("withdrawal_date")
        if withdrawal_date_val and isinstance(withdrawal_date_val, datetime.datetime):
            withdrawal_date_val = withdrawal_date_val.strftime("%Y-%m-%d")
        elif withdrawal_date_val and isinstance(withdrawal_date_val, str) and len(withdrawal_date_val) > 10:
            withdrawal_date_val = withdrawal_date_val[:10]
        ws.cell(row=r, column=2, value=withdrawal_date_val)

        date_val = receipt.get("date")
        if date_val and isinstance(date_val, datetime.datetime):
            date_val = date_val.strftime("%Y-%m-%d")
        elif date_val and isinstance(date_val, str) and len(date_val) > 10:
            date_val = date_val[:10]
        ws.cell(row=r, column=3, value=date_val)

        if receipt.get("type") == "增值税发票":
            ws.cell(row=r, column=4, value=receipt.get("receipt_number", ""))
        else:
            ws.cell(row=r, column=4, value="")

        desc = receipt.get("description", "")
        desc_cell = ws.cell(row=r, column=5, value=desc)
        ws.cell(row=r, column=6, value=receipt.get("person", ""))
        ws.cell(row=r, column=7, value=ev_no)

        major_cell = ws.cell(row=r, column=8, value=receipt.get("account_major", ""))
        minor_cell = ws.cell(row=r, column=9, value=receipt.get("account_minor", ""))
        
        if receipt.get("is_mapped") is False:
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            desc_cell.fill = red_fill
            major_cell.fill = red_fill
            minor_cell.fill = red_fill

        deposit_cny = receipt.get("deposit_cny") if not split_marker else None
        if deposit_cny is not None:
            ws.cell(row=r, column=10, value=deposit_cny)

        deposit_rate = receipt.get("deposit_rate") if not split_marker else None
        if deposit_rate is not None:
            ws.cell(row=r, column=11, value=deposit_rate)

        deposit_usd = receipt.get("deposit_usd") if not split_marker else None
        if deposit_usd is not None:
            ws.cell(row=r, column=12, value=deposit_usd)

        if rmb_amt is not None and receipt.get("type") != "입금영수증":
            ws.cell(row=r, column=13, value=rmb_amt)

        if rate_formula:
            ws.cell(row=r, column=14, value=rate_formula)

        if receipt.get("type") == "입금수수료" and receipt.get("withdrawal_usd") is not None and not split_marker:
            ws.cell(row=r, column=15, value=receipt.get("withdrawal_usd"))
            ws.cell(row=r, column=13, value=f"=O{r}*N{r}")
        elif receipt.get("type") == "입금영수증":
            ws.cell(row=r, column=15, value="")
        else:
            ws.cell(row=r, column=15, value=f"=M{r}/N{r}")

        ws.cell(row=r, column=16, value=f"=P{r-1}+J{r}-M{r}")
        ws.cell(row=r, column=17, value=f"=Q{r-1}+L{r}-O{r}")

        remark_text = receipt.get("remark") or ""
        if split_marker:
            remark_text = f"{remark_text} {split_marker}".strip()
        if remark_text:
            ws.cell(row=r, column=18, value=remark_text)

        warning = receipt.get("validation_warning")
        system_remarks = []
        if warning:
            if "✅" in warning:
                system_remarks.append(warning)
            else:
                system_remarks.append(f"[{warning}]")
            
        if receipt.get("type") == "增值税发票":
            if not receipt.get("tax_code_valid"):
                if receipt.get("description") == "통신비":
                    system_remarks.append("✅ 통신비 - 파표 회사코드 X")
                else:
                    system_remarks.append("파표 회사코드 불일치/누락")
        elif receipt.get("type") and receipt.get("type") != "银行回单":
            system_remarks.append("파표 외 영수증")
                
        if system_remarks:
            from openpyxl.styles import Font, Alignment
            remark_str = "\n".join(system_remarks)
            item_cell = ws.cell(row=r, column=19, value=remark_str)
            item_cell.alignment = Alignment(wrap_text=True)
            if "✅" not in remark_str:
                item_cell.font = Font(color="FF0000")

    current_row = 22
    for idx, receipt in enumerate(receipts):
        evidence_no = receipt.get("evidence_no", idx + 1)
        r_type = receipt.get("type")
        
        if r_type == '입금영수증':
            write_receipt_row(ws, current_row, receipt, evidence_no, None, None, "")
            current_row += 1
            continue

        if r_type == '입금수수료':
            # 전도금 수수료도 전월 최종 환산요율 적용
            write_receipt_row(ws, current_row, receipt, evidence_no, None, prev_rate_val, "")
            
            # 수수료만큼 이월 잔액(RMB) 차감
            if remaining_carry_rmb > 0.001:
                fee_usd = receipt.get('amount') or 0.0
                fee_rmb = fee_usd * prev_rate_val
                remaining_carry_rmb = max(0.0, remaining_carry_rmb - fee_rmb)
                
            current_row += 1
            continue

        receipt_rmb = receipt.get('amount') or 0.0
        
        if remaining_carry_rmb > 0.001:
            if receipt_rmb <= remaining_carry_rmb:
                write_receipt_row(ws, current_row, receipt, evidence_no, receipt_rmb, prev_rate_val, "")
                remaining_carry_rmb -= receipt_rmb
                current_row += 1
            else:
                write_receipt_row(ws, current_row, receipt, evidence_no, remaining_carry_rmb, prev_rate_val, "(분할 1/2)")
                current_row += 1
                
                rest_amount = receipt_rmb - remaining_carry_rmb
                remaining_carry_rmb = 0.0
                
                curr_rate_idx = get_latest_rate_row(receipt.get('withdrawal_date'))
                curr_rate_form = f"='통장내역(환율표)'!D{curr_rate_idx}" if curr_rate_idx else ""
                write_receipt_row(ws, current_row, receipt, evidence_no, rest_amount, curr_rate_form, "(분할 2/2)")
                current_row += 1
        else:
            curr_rate_idx = get_latest_rate_row(receipt.get('withdrawal_date'))
            curr_rate_form = f"='통장내역(환율표)'!D{curr_rate_idx}" if curr_rate_idx else ""
            write_receipt_row(ws, current_row, receipt, evidence_no, receipt_rmb, curr_rate_form, "")
            current_row += 1

    wb.save(output_path)
    return output_path



def get_available_exports() -> list:
    """생성된 엑셀 파일 목록 반환"""
    ensure_output_dir()
    files = []
    for f in os.listdir(OUTPUT_DIR):
        if f.endswith(".xlsx"):
            filepath = os.path.join(OUTPUT_DIR, f)
            files.append({
                "filename": f,
                "path": filepath,
                "size": os.path.getsize(filepath),
                "created": datetime.datetime.fromtimestamp(
                    os.path.getctime(filepath)
                ).isoformat()
            })
    files.sort(key=lambda x: x["created"], reverse=True)
    return files
